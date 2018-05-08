'''
Program to take a source SVG file, find the placeholder text 'Participant'
and replace it with each participant's name which is read from an Excel 
document having the full name in column specified
'''

import os
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
mypath = 'C:/Users/omahmed/Desktop/'
textToSearch = 'Participant'
# Read in the source file
with open(mypath + 'sourcefile.svg', 'r') as file :
    filedata = file.read()
# Load the Excel sheet with all the participant names
wb = load_workbook(mypath + 'sourcesheet.xlsx')
sheet = wb['Participants']
namesColumn = 1 # The column in which the names of the participants are stored 
for i in range(2, sheet.max_row):  # Ignore header row and go up to the max row
    currVal = sheet.cell(row=i, column=namesColumn).value
    if currVal != None: # Cell is not empty
        name = currVal.title()  # Change name to Title case
        print("Currently processing:", name)
        # Replace the target string
        currfile = filedata.replace(textToSearch, name)
        # Write the file out
        with open(mypath + 'certificates/' + name + '.svg', 'w') as file:
            file.write(currfile)
        os.chdir(mypath + 'certificates')
        os.system('inkscape --without-gui --export-pdf="export/'+ name + '.pdf" "'+ name + '.svg" --export-area-page')
